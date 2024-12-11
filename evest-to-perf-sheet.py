#!/usr/bin/env python
# coding: utf-8

# In[1]:


# Where your data is stored
data_loc = "##############################"


# In[2]:


# For DataFrames
import pandas as pd

# For Excel Editing
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import Workbook, load_workbook

# To Get the last day of the previous month
from datetime import date, datetime, timedelta

# To check if a file already exists
import os

# For copying the spreadsheet to another spreadsheet
from copy import copy


# In[ ]:


def EOPM():
    """
    Gets the last day of the previous month.

    This function calculates the last day of the previous month based on the current date.

    Returns:
        last_day_current_month (datetime.datetime): The last day of the previous month.
    """
    # Gets current date
    today = datetime.today()

    # Gets first day of the current month
    first_day_this_month = datetime(today.year, today.month, 1)

    # Subtracts one day from the first day of this month to get the last day of the previous month
    last_day_prev_month = first_day_this_month - timedelta(days=1)

    return last_day_prev_month

def insert_cash_row(df):
    """
    Inserts a row for cash holdings in a DataFrame.

    This function identifies the row in the DataFrame where the 'Identifier' column is NaN (representing cash), 
    creates a new row for cash holdings with specific values, removes the original cash row, and 
    appends the new row to the DataFrame. The DataFrame is then sorted by the 'Security Name' column.

    Args:
        df (pandas.DataFrame): The DataFrame containing the holdings data, including a row for cash with NaN 'Identifier'.

    Returns:
        df (pandas.DataFrame): The updated DataFrame with the new cash row inserted and sorted by 'Security Name'.
    """
    
    idx = df['Identifier'].isna().idxmax()
    cash = df.iloc[idx]
    new_row = ['CASH', 'CASH', 'USD', 'US DOLLAR', 'Cash', cash['# of Shares'],
           cash['Security Price'], cash['Weight (%)'], 'United States', cash['Market Value']]
    df.drop(idx, inplace=True)
    df = df.reset_index(drop = True)
    df.loc[len(df)] = new_row
    df = df.sort_values(by='Security Name')


def get_securities(holds, df):
    """
    Extracts securities and their associated countries from the holdings DataFrame.

    This function processes a holdings DataFrame to identify rows representing securities 
    (determined by non-null values in the 3rd column). It backtracks from each security 
    row to find the most recent country (determined by non-null values in the 2nd column) 
    and compiles this information into a new DataFrame.

    Args:
        holds (pandas.DataFrame): DataFrame containing various data including securities and country information.
        df (pandas.DataFrame): DataFrame that will be populated with the securities and their countries.

    Returns:
        df (pandas.DataFrame): DataFrame that contains all the securities with their respective countries.
    """
    
    for index, row in holds.iterrows():
    
    # Only if it's an actual security
        if pd.notna(row['Unnamed: 2']):
        
            # Getting Country
            country_index = index - 1
            while holds.isna().iloc[country_index]['Unnamed: 1']:
                country_index -= 1
            country = holds.iloc[country_index]['Unnamed: 1']
    
            # New row in new dataframe
            df.loc[len(df)] = [row['ISIN'], 'ISIN', row['Ticker'], row['Unnamed: 2'], 'Common Stock',
                               row['Pos'], row['Px Close'], row['% Wgt'], country.upper(), row['Mkt Val']]
            
    return df

def extract_perf_info(perf, df):
    """
    Extracts performance information for specific strategies and append it to a DataFrame.

    This function iterates over a list of predefined strategies, extracts their gross and net performance 
    data from the provided DataFrame, and appends this information to another DataFrame with additional 
    metadata such as the strategy name and the current date.

    Args:
        perf (pandas.DataFrame): The DataFrame containing performance data with a specific structure.
        df (pandas.DataFrame): The DataFrame to which the extracted performance information will be appended.

    Returns:
        df (pandas.DataFrame): The updated DataFrame with performance information appended and indexed by strategy name.
    """
    # Iterates over each strategy 
    strategies = ["EAFE Small Cap Value", "EM Small Cap Value", "Int'l Small Cap Value", "ISC Impact"]
    renames = ["EAFE", "EM", "ISC Composite", "ISCIO"]
    for strat, name in zip(strategies, renames):
    
        # Searches the dataframe for the strategy
        mask = perf['Unnamed: 1'].str.contains(strat, na=False)
        result = perf[mask]
        gross = list(perf.iloc[result.index + 2]['Unnamed: 2'])[0]
        net = list(perf.iloc[result.index + 2]['Unnamed: 3'])[0]
        row = {'Gross': gross, 'Net': net, 'Strategy': name, 'Date': date.today().strftime("%m/%d/%Y")}
    
        # Adds the strategy's performance to the dataframe
        df.loc[len(df)] = [gross, net, name, date.today().strftime("%m/%d/%Y")]
    
    # Sets the index of the rows to the strategy and returns the dataframe
    df.set_index('Strategy', inplace=True)
    return df

def get_countries_weighted(holds):
    """
    Calculates weighted percentages for countries based on holdings data.

    This function calculates the weighted percentages for each country based on the provided holdings DataFrame.
    It first determines the weight of cash in the holdings, then calculates the weight for each country excluding cash.
    
    Note that 'Unnamed: 1' is for country names and '% Wgt' is for weights, and that Cash is identified by rows
    containing 'US Dollar Spot', and its weight is used to calculate country weights. Also, the returned dictionary
    excludes the 'Not Classified' category (which is USD).
    
    Args:
        holds (pandas.DataFrame): DataFrame containing holdings data including countries and their weights.

    Returns:
        countries (dict): A dictionary where keys are country names and values are their weighted percentages.
    """
    # Getting Cash/Weight Makeup using boolean mask on DataFrame
    mask = holds['Unnamed: 2'].str.contains('US Dollar Spot', na=False)
    result = holds[mask]
    cash = list(holds.iloc[result.index]['% Wgt'])[0]
    weight = 1 / ((100 - cash) * 0.01)

    # Getting countries
    countries = {}
    for index, row in holds.iterrows():
        
        # Only if it's a country, not security
        if pd.notna(row['Unnamed: 1']):
            countries[row['Unnamed: 1']] = round(row['% Wgt'] * weight, 2)
    
    del countries['Not Classified']
    return countries

def update_alloc(countries, alloc):
    """
    Updates the 'Country (%)' column in a DataFrame containing allocation data based on country names and their percentages.
    
    Iterates through each country in the 'countries' dictionary (key: country name, value: percentage).
    For each country, it searches for its presence in the 'Market' column of the DataFrame 'alloc' and updates the
    corresponding 'Country (%)' value if the country is found and the 'Country (%)' value is not already present.

    Parameters:
        alloc (pd.DataFrame): DataFrame containing allocation data with columns like 'Market' and 'Country (%)'.

    Returns:
        alloc (pd.DataFrame): Updated DataFrame with 'Country (%)' column values updated based on the specified countries and their percentages.
    """
    titles = ['North America', 'United Kingdom', 'Euroland (EU) Countries', 'Non-Euroland (EU) Countries',
              'Far East & Australasia', 'Other', 'Latin America', 'Africa/Middle East', 'Eastern Europe',
              'Far East ex-China', 'China', 'Other Emerging Markets', 'Emerging Market Total']
    for country in countries.items():
        mask = (alloc['Market'].str.contains(country[0], na=False, regex=False) & ~alloc['Market'].isin(titles))
        result = alloc[mask]
        if (pd.notna(alloc.loc[result.index, 'Country (%)']).any()):
            alloc.loc[result.index[0], 'Country (%)'] = country[1]
    return alloc

def create_holds_excel(df, file_name):
    """
    Creates an Excel file with formatted holdings data.

    This function writes a DataFrame to an Excel file, applying specific formatting 
    to ensure column names are visible and have consistent styling. The file is appended 
    to an existing Excel workbook using the 'openpyxl' engine.

    Args:
        df (pandas.DataFrame): The DataFrame containing the holdings data to be written to the Excel file.
        file_name (str): The name of the Excel file where the data will be written. It should include the path and extension.

    Returns:
        None
    """
    try:
        sheet = 'Holdings'
        
        if not os.path.exists(data_loc + file_name):
            df.to_excel(data_loc + file_name, index=False, sheet_name=sheet, header=True)
        
        # Formatting so that the names are all visible and columns names have the same formatting
        writer = pd.ExcelWriter(data_loc + file_name, engine='openpyxl', mode='a', if_sheet_exists='replace')
        df.to_excel(writer, index=False, sheet_name=sheet, header=True)
        worksheet = writer.sheets[sheet]
        
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['I'].width = 20
        
        for cell in worksheet[1]:  # Access the first row (header row)
            cell.font = Font(name='Arial', size=8, underline='single', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='bottom')
            cell.border = Border(top=None, bottom=None, left=None, right=None)
        
        writer.close()
        
    except PermissionError:
        print(f"Permission denied: The file {file_name} is open. Please close the file and try again.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        writer.close()

def create_perf_excel(df, file_name):
    """
    Creates an Excel file with performance data and apply specific formatting.

    This function writes a DataFrame to an Excel file, appending to an existing workbook. 
    It formats the columns for better readability and applies a percentage format to specified columns.
    Columns 'A' to 'D' are resized for better visibility, and columns 'B' and 'C' are formatted
    to display percentages.

    Args:
        df (pandas.DataFrame): The DataFrame containing the performance data to be written to the Excel file.
        file_name (str): The name of the Excel file where the data will be written. It should include the path and extension.

    Returns:
        None
    """
    try:
        sheet = 'Performance'
        
        if not os.path.exists(data_loc + file_name):
            df.to_excel(data_loc + file_name, index=False, sheet_name=sheet)
            
        # Creates a Pandas Excel writer using XlsxWriter as the engine
        writer = pd.ExcelWriter(data_loc + file_name, engine='openpyxl', mode='a', if_sheet_exists='replace')
        
        # Converts dataframe to an XlsxWriter Excel object
        df.to_excel(writer, sheet_name=sheet, header=True)
        
        worksheet = writer.sheets['Performance']
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 15
        
        for col in ['B', 'C']:
           for cell in worksheet[col]:
               cell.number_format = '0.00%'
    
        writer.close()
    except PermissionError:
        print(f"Permission denied: The file {file_name} is open. Please close the file and try again.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        writer.close()

def create_alloc_excel(alloc, file_name):
    """
    Creates an Excel file with allocation data and apply specific formatting.

    This function writes an allocation DataFrame to an Excel file, creating or updating the 'Allocations' sheet.
    It applies various formatting styles such as font, column width, number format, cell merging, and background colors
    to improve readability and organization of the data.

    Args:
        alloc (pandas.DataFrame): The DataFrame containing allocation data to be written to the Excel file.
        file_name (str): The name of the Excel file where the data will be written. It should include the path and extension.

    Returns:
        None
    """
    try:
        sheet = 'Allocations'
        
        if not os.path.exists(data_loc + file_name):
            df.to_excel(data_loc + file_name, index=False, sheet_name=sheet)
            
        # Creates a Pandas Excel writer using XlsxWriter as the engine
        writer = pd.ExcelWriter(data_loc + file_name, engine='openpyxl', mode='a', if_sheet_exists='replace')
        
        # Converts dataframe to an XlsxWriter Excel object
        df.to_excel(writer, sheet_name=sheet, header=True)
        
        worksheet = writer.sheets[sheet]
        
        # Main Font for the spreadsheet
        font_style = Font(name='Tahoma', size=10)
        
        # Iterate through all cells in the sheet and apply the font settings
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font_style
        
        # Formatting Column Width
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['C'].width = 30

        # Populates sheet
        for r_idx, row in enumerate(alloc.itertuples(), start=1):
            for c_idx, value in enumerate(row[1:], start=1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Formatting numbers
        for col in ['B', 'C']:
            for cell in worksheet[col]:
                cell.number_format = '0.00'
        
        # Making Title rows Gray, Bolded, and Centered
        titles = ['North America', 'United Kingdom', 'Euroland (EU) Countries', 'Non-Euroland (EU) Countries',
              'Far East & Australasia', 'Other', 'Latin America', 'Africa/Middle East', 'Eastern Europe',
              'Far East ex-China', 'China', 'Other Emerging Markets', 'Emerging Market Total']
        
        gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        bold_font = Font(name='Tahoma', bold=True, size=10)
        for row in worksheet.iter_rows(min_row=1, max_row=len(alloc), min_col=1, max_col=len(alloc.columns)):
            for cell in row:
                if cell.value in titles:
                    for c in row:
                        c.fill = gray_fill
                        c.font = bold_font
                        c.alignment = Alignment(horizontal='center', vertical='center')
                    break  # Once the row is colored, no need to check further cells in the row
        
        # Making 2nd and 3rd columns always centered (text-align)
        for col in ['B', 'C']:
            for cell in worksheet[col]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formatting first specific columns
        light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        for cell_ref in ['A1', 'A2', 'A3', 'A4', 'A5', 'A6']:
            cell = worksheet[cell_ref]
            cell.fill = light_blue_fill
        
        worksheet['A10'].font = bold_font
        worksheet['A14'].font = bold_font
        worksheet['A12'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['A12'].font = Font(name='Tahoma', size=11, bold=True, color='00008B')
        
        worksheet.insert_rows(1)  # To line it up with the template
        
        worksheet.merge_cells('A1:K1')
        worksheet.merge_cells('A2:K2')
        worksheet.merge_cells('A3:K3')
        worksheet.merge_cells('A4:K4')
        worksheet.merge_cells('A5:K5')
        worksheet.merge_cells('A6:K6')
        worksheet.merge_cells('A7:K7')
        worksheet.merge_cells('A8:K8')
        worksheet.merge_cells('A13:C13')
        
        writer.close()

    except PermissionError:
        print(f"Permission denied: The file {file_name} is open. Please close the file and try again.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        writer.close()

def get_row(phrase, column, sheet):
    """
    Searches for a phrase in a specific column of an Excel sheet and returns the row number where it is found.

    Args:
        phrase (str): The phrase to search for.
        column (str): The column letter where the search is performed.
        sheet (Worksheet): The worksheet object to search in.

    Returns:
        row (int): The row number where the phrase is found. Returns -1 if the phrase is not found.
    """
    for row in range(1, sheet.max_row + 1):
        cell_val = sheet[f'{column}{row}'].value
        if cell_val == phrase:
            return row
    return -1

def get_num_entries(column, sheet):
    """
    Counts the number of non-empty entries in a specific column of an Excel sheet, starting from row 9.

    Args:
        column (str): The column letter to count entries in.
        sheet (Worksheet): The worksheet object to count entries in.

    Returns:
        count (int): The number of non-empty entries in the specified column.
    """
    count = 0
    for row in range(9, sheet.max_row + 1):
        cell_val = sheet[f'{column}{row}'].value
        if cell_val is not None:
            count += 1
    return count

def get_num_countries(chars):
    """
    Counts the number of non-empty entries in column 'B' and ensures there is a
    corresponding non-empty entry in column 'D' in the 'chars' worksheet. This 
    ensures that countries with no securities are not counted.

    Returns:
        count (int): The count of entries meeting the criteria.
    """
    count = 0
    for row in range(9, chars.max_row + 1):
        cell_val = chars[f'{"B"}{row}'].value
        check = chars[f'{"D"}{row}'].value
        if cell_val is not None and check is not None:
            count += 1
    return count

def get_top_10_sum(chars):
    """
    Calculates the sum of the top 10 highest values in column 'D' (weights) starting from row 14,
    ignoring rows where column 'B' (countries) is not empty. Rows with column 'B' must be empty to
    ensure this is a security, not a country.

    Returns:
        sum_top_10 (float): The sum of the top 10 highest values.
    """
    values = []
    for row in range(14, chars.max_row + 1):
        cell_val = chars[f'{"D"}{row}'].value
        ignore = chars[f'{"B"}{row}'].value
        if cell_val is not None and ignore is None:
            values.append(cell_val)

    sorted_vals = sorted(values, reverse=True)
    top_10 = sorted_vals[:10]
    sum_top_10 = sum(top_10)
    return sum_top_10

def create_chars_excel(wb, wb2):
    """
    Updates the 'CharacteristicsUpdated' sheet (the formatted sheet) in the
    workbook with various values from 'Characteristics' and 'Sectors' sheets.

    Args:
        wb (Workbook): The openpyxl Workbook object containing the sheets.

    Returns:
        format (Worksheet): The updated 'CharacteristicsUpdated' sheet object.
    """
    format = wb['CharacteristicsUpdated']
    chars = wb['Characteristics']
    sectors = wb['Sectors']
    overall_row = get_row('##############################', 'A', chars)

    format['B17'] = chars['D' + str(get_row('US Dollar Spot', 'C', chars))].value / 100
    format['B19'] = get_num_entries('C', chars)
    format['B20'] = get_top_10_sum(chars) / 100
    format['B21'] = get_num_countries(chars)
    format['B24'] = chars['AC' + str(overall_row)].value / 100
    format['B28'] = chars[f'{"U"}{overall_row}'].value / 100
    format['B29'] = chars[f'{"K"}{overall_row}'].value
    format['B30'] = chars[f'{"M"}{overall_row}'].value
    format['B31'] = chars[f'{"O"}{overall_row}'].value
    format['B32'] = chars[f'{"Q"}{overall_row}'].value
    format['B33'] = chars[f'{"S"}{overall_row}'].value
    format['B34'] = chars[f'{"AA"}{overall_row}'].value / 100
    format['B37'] = chars[f'{"W"}{overall_row}'].value / 100
    
    format['B41'] = chars[f'{"G"}{overall_row}'].value
    format['B42'] = chars[f'{"I"}{overall_row}'].value
    
    format['B48'] = (100 - chars['D' + str(get_row('US Dollar Spot', 'C', chars))].value) / 100
    format['B49'] = 0
    format['B50'] = 0
    format['B51'] = format['B17'].value
    format['B52'] = 0

    format['B55'] = 0
    format['B56'] = 0
    holds = wb2['Holdings']
    weight = 1 / (1 - format['B51'].value)
    format['B57'] = holds['D' + str(get_row('7.5-15B', 'B', holds))].value * weight / 100
    format['B58'] = holds['D' + str(get_row('1.5-7.5B', 'B', holds))].value * weight / 100
    format['B59'] = holds['D' + str(get_row('750M-1.5B', 'B', holds))].value * weight / 100
    format['B60'] = holds['D' + str(get_row('400-750M', 'B', holds))].value * weight / 100
    format['B61'] = holds['D' + str(get_row('<400M', 'B', holds))].value * weight / 100
    
    format['B64'] = sectors['E' + str(get_row('Communication Services', 'B', sectors))].value
    format['B65'] = sectors['E' + str(get_row('Consumer Discretionary', 'B', sectors))].value
    format['B66'] = sectors['E' + str(get_row('Consumer Staples', 'B', sectors))].value
    format['B67'] = sectors['E' + str(get_row('Energy', 'B', sectors))].value
    format['B68'] = sectors['E' + str(get_row('Financials', 'B', sectors))].value
    format['B69'] = sectors['E' + str(get_row('Health Care', 'B', sectors))].value
    format['B70'] = sectors['E' + str(get_row('Industrials', 'B', sectors))].value
    format['B71'] = sectors['E' + str(get_row('Information Technology', 'B', sectors))].value
    format['B72'] = sectors['E' + str(get_row('Materials', 'B', sectors))].value
    format['B73'] = sectors['E' + str(get_row('Real Estate', 'B', sectors))].value
    format['B74'] = sectors['E' + str(get_row('Utilities', 'B', sectors))].value
    format['B75'] = 0
    
    for i in range(64, 76):
        if type(format['B' + str(i)].value) == type(None):
            format['B' + str(i)] = 0
        else:
            format['B' + str(i)] = format['B' + str(i)].value / 100

    format['B78'] = format['B71'].value
    format['B79'] = format['B69'].value
    format['B80'] = format['B65'].value
    format['B81'] = format['B66'].value
    format['B82'] = format['B70'].value
    format['B83'] = format['B72'].value
    format['B84'] = format['B68'].value
    format['B85'] = format['B67'].value
    format['B86'] = format['B74'].value
    format['B87'] = format['B64'].value
    format['B88'] = format['B73'].value
    
    format['A14'] = '##############################'
    format['A14'].font = Font(name='Tahoma', size=12, bold=True)
    format['A14'].alignment = Alignment(horizontal='center', vertical='center')

    return format

def copy_sheet_attributes(source_sheet, target_sheet):
    """
    Copies various attributes from a source worksheet to a target worksheet. 
    Specifically, these are the sheet format, properties, merged cells, page margins, freeze panes, row dimensions,
    specific column widths, and hidden properties settings from the source sheet.

    Parameters:
    - source_sheet (Worksheet): The source worksheet from which attributes are copied.
    - target_sheet (Worksheet): The target worksheet to which attributes are copied.

    Returns:
    - None
    """
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is not None:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width)
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

def copy_cells(source_sheet, target_sheet):
    """
    Copies cell values, styles, hyperlinks, and comments from a source worksheet to a target worksheet,
    excluding cells listed in the 'forbidden' list. The 'forbidden' list contains merged cells that are
    read-only and will throw an error if you try reassigning them.

    Parameters:
    - source_sheet (Worksheet): The worksheet from which cells are copied.
    - target_sheet (Worksheet): The worksheet to which cells are copied.

    Returns:
    - None
    """
    forbidden = ['G12', 'H12', 'F13', 'B14', 'C14', 'D14', 'E14', 'B63', 'B77']
    for row in source_sheet.iter_rows():
        for source_cell in row:
            
            if source_cell.coordinate in forbidden:
                continue # Skip cells listed in forbidden

            # Get corresponding cell in target sheet
            target_cell = target_sheet.cell(row=source_cell.row, column=source_cell.col_idx)

            # Copy value
            target_cell.value = source_cell.value

            # Copy style if present
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            # Copy hyperlink if present
            if source_cell.hyperlink:
                target_cell.hyperlink = copy(source_cell.hyperlink)
                
            # Copy comment if present
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

def copy_chars_sheet_to_main(sheet, file_name):
    """
    Copy a sheet to another workbook, ensuring the target workbook has a sheet named 'Characteristics'.

    Args:
        sheet (Worksheet): The openpyxl Worksheet object to copy.
        file_name (str): The name of the target Excel file.

    Returns:
        None
    """
    sheet_name = 'Characteristics'
    
    if not os.path.exists(data_loc + file_name):
        target = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        workbook.save(data_loc + file_name)
    else:
        target = load_workbook(data_loc + file_name)

    if sheet_name in target.sheetnames:
        # Get the specific sheet if it exists
        target_sheet = target[sheet_name]
    else:
        # Create a new sheet if it doesn't exist
        target_sheet = target.create_sheet(title=sheet_name)
    
    copy_cells(sheet, target_sheet)  # copy all the cell values and styles
    copy_sheet_attributes(sheet, target_sheet)
    
    target.save(data_loc + file_name)


# In[4]:


# Holdings
holds = pd.read_excel(data_loc + "##############################", skiprows=10)
df = pd.DataFrame(columns = ['Identifier', 'Identifier Type', 'Ticker', 'Security Name', 'Security Type',
                             '# of Shares', 'Security Price', 'Weight (%)', 'Country', 'Market Value'])

df = get_securities(holds, df)
insert_cash_row(df)
create_holds_excel(df, "##############################")


# In[5]:


# Performance
perf = pd.read_excel(data_loc + "##############################" + EOPM().strftime("%Y-%m-%d") + '.xlsx')
perf.drop(inplace=True, columns=['Unnamed: 0'])
df = pd.DataFrame(columns=['Gross', 'Net', 'Strategy', 'Date'])


df = extract_perf_info(perf, df)
create_perf_excel(df, "##############################")


# In[6]:


# Allocations
countries = get_countries_weighted(holds)

alloc = pd.read_excel(data_loc + '##############################')
alloc.columns = ['Market', 'Country (%)', 'Currency (%)']

alloc = update_alloc(countries, alloc)
create_alloc_excel(alloc, "##############################")


# In[7]:


# Characteristics
try:
    wb = load_workbook(data_loc + '##############################')
    wb2 = load_workbook(data_loc + '##############################')
    format = create_chars_excel(wb, wb2)
    copy_chars_sheet_to_main(format, "##############################")
except PermissionError:
    print(f"Permission denied: The file {'##############################'} is open. Please close the file and try again.")
except Exception as e:
    print(f"An unexpected error occurred: {e}")
